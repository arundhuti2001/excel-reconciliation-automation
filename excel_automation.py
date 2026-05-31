from openpyxl import load_workbook
from difflib import SequenceMatcher
import os
import copy

# =========================================================
# FILE PATHS
# =========================================================

BASE_DIR = r"c:\Users\arund\Downloads\Arundhuti_Dey_Interview Second Round\Interview Second Round\Project 2\project_2_excel_automation"

INPUT_FILE = os.path.join(BASE_DIR, "input", "Input_file.xlsx")
MAPPING_FILE = os.path.join(BASE_DIR, "input", "Mapping_File.xlsx")
OUTPUT_TEMPLATE = os.path.join(BASE_DIR, "input", "Output_File.xlsx")

FINAL_OUTPUT = os.path.join(BASE_DIR, "output", "final_output.xlsx")
LOG_FILE = os.path.join(BASE_DIR, "match_log.txt")

# =========================================================
# TARGET SHEETS
# =========================================================

TARGET_SHEETS = [
    "Mr. Ajay Tyagi",
    "Mr. Amit Premchandani",
    "Mr. Deepesh Agarwal",
    "Mr. Sachin Trivedi",
    "Mr. Kamal Gada",
    "Mr. Karthikraj Lakshmanan",
    "Mr. Nitin Jain",
    "Mr. Ravi Gupta",
    "Mr. V Srivatsa",
    "Mr. Vicky Punjabi",
    "Mr. Vishal Chopda",
    "Mr. Amit Sharma",
    "Mr. Anurag Mittal",
    "Mr. Ayush Jain",
    "Mr. Jaydeep Bhowal",
    "Mr. Pankaj Pathak",
    "Mr. Sharwan Kumar Goyal",
    "Mr. Sunil Patil"
]

# =========================================================
# HELPER FUNCTIONS
# =========================================================

def similarity(a, b):

    return SequenceMatcher(
        None,
        str(a).lower(),
        str(b).lower()
    ).ratio() * 100


def clean_text(value):

    if value is None:
        return ""

    return str(value).strip().lower()


def copy_style(source_cell, target_cell):

    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)

    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)

    if source_cell.fill:
        target_cell.fill = copy.copy(source_cell.fill)

    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)

    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)

    if source_cell.number_format:
        target_cell.number_format = copy.copy(source_cell.number_format)

    if source_cell.protection:
        target_cell.protection = copy.copy(source_cell.protection)

# =========================================================
# LOAD FILES
# =========================================================

print("\nLoading Workbooks...")

input_wb = load_workbook(INPUT_FILE, data_only=True)

mapping_wb = load_workbook(MAPPING_FILE, data_only=True)

# IMPORTANT:
# LOAD OUTPUT WITHOUT data_only
# TO KEEP SAME FORMAT/LAYOUT/COLORS
output_wb = load_workbook(OUTPUT_TEMPLATE)

print("Workbooks Loaded Successfully!")

# =========================================================
# CREATE LOG FILE
# =========================================================

log = open(LOG_FILE, "w", encoding="utf-8")

# =========================================================
# READ INPUT FILE
# =========================================================

print("\nReading Input Data...")

input_records = []

for sheet in input_wb.worksheets:

    for row in sheet.iter_rows():

        if len(row) == 0:
            continue

        scheme_name = row[0].value

        if scheme_name is None:
            continue

        scheme_name = str(scheme_name).strip()

        if scheme_name == "":
            continue

        if "scheme" in scheme_name.lower():
            continue

        row_values = []

        for cell in row:
            row_values.append(cell.value)

        input_records.append({
            "scheme": scheme_name,
            "values": row_values
        })

print(f"Total Input Records : {len(input_records)}")

# =========================================================
# READ MAPPING FILE
# =========================================================

print("\nReading Mapping Data...")

mapping_records = []

for sheet in mapping_wb.worksheets:

    for row in sheet.iter_rows():

        values = [cell.value for cell in row]

        if len(values) < 2:
            continue

        scheme_name = values[0]

        if scheme_name is None:
            continue

        scheme_name = str(scheme_name).strip()

        if scheme_name == "":
            continue

        if "scheme" in scheme_name.lower():
            continue

        mapping_records.append({
            "scheme": scheme_name,
            "values": values
        })

print(f"Total Mapping Records : {len(mapping_records)}")

# =========================================================
# PROCESS OUTPUT SHEETS
# =========================================================

print("\nUpdating Output Workbook...")

for sheet_name in TARGET_SHEETS:

    if sheet_name not in output_wb.sheetnames:
        continue

    ws = output_wb[sheet_name]

    print(f"\nProcessing Sheet : {sheet_name}")

    max_row = ws.max_row

    row = 1

    while row <= max_row:

        cell_value = ws[f"A{row}"].value

        if cell_value is None:
            row += 1
            continue

        scheme_name = str(cell_value).strip()

        if scheme_name == "":
            row += 1
            continue

        if "scheme" in scheme_name.lower():
            row += 1
            continue

        # =====================================================
        # FIND BEST INPUT MATCH
        # =====================================================

        best_match = None
        best_score = 0

        for record in input_records:

            score = similarity(
                scheme_name,
                record["scheme"]
            )

            if score > best_score:
                best_score = score
                best_match = record

        # =====================================================
        # FIND BEST MAPPING MATCH
        # =====================================================

        best_mapping_match = None
        best_mapping_score = 0

        for mapping in mapping_records:

            score = similarity(
                scheme_name,
                mapping["scheme"]
            )

            if score > best_mapping_score:
                best_mapping_score = score
                best_mapping_match = mapping

        # =====================================================
        # PRINT LOGS
        # =====================================================

        print("\n================================")
        print(f"SHEET : {sheet_name}")
        print(f"ROW : {row}")
        print(f"OUTPUT SCHEME : {scheme_name}")

        log.write("\n================================\n")
        log.write(f"SHEET : {sheet_name}\n")
        log.write(f"ROW : {row}\n")
        log.write(f"OUTPUT SCHEME : {scheme_name}\n")

        # =====================================================
        # UPDATE USING INPUT FILE
        # KEEP SAME FORMAT / COLORS / MERGED CELLS
        # ONLY UPDATE VALUES
        # =====================================================

        if best_match and best_score >= 80:

            print(f"INPUT MATCH : {best_match['scheme']}")
            print(f"INPUT SCORE : {best_score}")

            log.write(f"INPUT MATCH : {best_match['scheme']}\n")
            log.write(f"INPUT SCORE : {best_score}\n")

            input_values = best_match["values"]

            # =================================================
            # COLUMN STRUCTURE
            #
            # A = Scheme Name
            # B = Inception Date
            # C = Managing Since
            # D = Benchmark
            # E = Additional Benchmark
            #
            # F = 1Y
            # G = 1Y Benchmark
            # H = 1Y Additional Benchmark
            #
            # I = 3Y
            # J = 3Y Benchmark
            # K = 3Y Additional Benchmark
            #
            # L = 5Y
            # M = 5Y Benchmark
            # N = 5Y Additional Benchmark
            # =================================================

            COLUMN_MAP = {
                0: "A",
                1: "B",
                2: "C",
                3: "D",
                4: "E",
                5: "F",
                6: "G",
                7: "H",
                8: "I",
                9: "J",
                10: "K",
                11: "L",
                12: "M",
                13: "N"
            }

            for input_index, excel_col in COLUMN_MAP.items():

                if input_index >= len(input_values):
                    continue

                new_value = input_values[input_index]

                if new_value is None:
                    continue

                target_cell = ws[f"{excel_col}{row}"]

                current_value = target_cell.value

                # =================================================
                # ALWAYS UPDATE PERFORMANCE COLUMNS
                # =================================================

                if excel_col in [
                    "F", "G", "H",
                    "I", "J", "K",
                    "L", "M", "N"
                ]:

                    target_cell.value = new_value

                else:

                    # ONLY FILL EMPTY CELLS
                    if (
                        current_value is None or
                        str(current_value).strip() == ""
                    ):

                        target_cell.value = new_value

        # =====================================================
        # UPDATE USING MAPPING FILE
        # =====================================================

        if best_mapping_match and best_mapping_score >= 80:

            print(f"MAPPING MATCH : {best_mapping_match['scheme']}")
            print(f"MAPPING SCORE : {best_mapping_score}")

            log.write(f"MAPPING MATCH : {best_mapping_match['scheme']}\n")
            log.write(f"MAPPING SCORE : {best_mapping_score}\n")

            mapping_values = best_mapping_match["values"]

            # =====================================================
            # MAPPING FILE STRUCTURE
            #
            # 0 = Scheme Name
            #
            # 1 = 1 Year
            # 2 = 1Y Benchmark
            # 3 = 1Y Additional Benchmark
            #
            # 4 = 3 Year
            # 5 = 3Y Benchmark
            # 6 = 3Y Additional Benchmark
            #
            # 7 = 5 Year
            # 8 = 5Y Benchmark
            # 9 = 5Y Additional Benchmark
            # =====================================================

            # =========================
            # 1 YEAR
            # =========================

            if len(mapping_values) > 1:
                ws[f"F{row}"] = mapping_values[1]

            if len(mapping_values) > 2:
                ws[f"G{row}"] = mapping_values[2]

            if len(mapping_values) > 3:
                ws[f"H{row}"] = mapping_values[3]

            # =========================
            # 3 YEAR
            # =========================

            if len(mapping_values) > 4:
                ws[f"I{row}"] = mapping_values[4]

            if len(mapping_values) > 5:
                ws[f"J{row}"] = mapping_values[5]

            if len(mapping_values) > 6:
                ws[f"K{row}"] = mapping_values[6]

            # =========================
            # 5 YEAR
            # =========================

            if len(mapping_values) > 7:
                ws[f"L{row}"] = mapping_values[7]

            if len(mapping_values) > 8:
                ws[f"M{row}"] = mapping_values[8]

            if len(mapping_values) > 9:
                ws[f"N{row}"] = mapping_values[9]

            print("1Y / 3Y / 5Y DATA UPDATED")

            log.write("1Y / 3Y / 5Y DATA UPDATED\n")

        row += 1

# =========================================================
# SAVE OUTPUT
# =========================================================

print("\nSaving Final Output...")

try:

    output_wb.save(FINAL_OUTPUT)

    print("\n================================")
    print("SUCCESS!")
    print("================================")

    print(f"\nFinal Output Saved At:\n{FINAL_OUTPUT}")

except PermissionError:

    print("\nERROR:")
    print("Close final_output.xlsx before running script.")

log.close()

print(f"\nLog File Saved At:\n{LOG_FILE}")